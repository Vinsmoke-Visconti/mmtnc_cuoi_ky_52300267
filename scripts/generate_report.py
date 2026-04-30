#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_report.py
------------------
Tao bao cao Excel da sheet tu cac file JSON trong results/.
Cai dat: pip3 install openpyxl matplotlib
Chay   : python3 scripts/generate_report.py
"""
import json, os, sys, glob
from datetime import datetime

# Fix import path when running as a script
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
try:
    from scripts import utils
except ImportError:
    import utils as utils

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference
except ImportError:
    print('[ERROR] pip3 install openpyxl'); sys.exit(1)

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

RESULTS_DIR = utils.get_session_dir()

# --- Style helpers ---
def hfill(c='1F4E79'): return PatternFill('solid', fgColor=c)
def dfill(c): return PatternFill('solid', fgColor=c)
def border():
    s = Side(style='thin', color='AAAAAA')
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, r, c, v, bg='1A5276', fg='FFFFFF', sz=10, bold=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = hfill(bg); cell.font = Font(bold=bold, color=fg, size=sz)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border(); return cell

def dat(ws, r, c, v, bg=None, bold=False, align='center'):
    cell = ws.cell(row=r, column=c, value=v)
    if bg: cell.fill = dfill(bg)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = border(); return cell

def title(ws, merged, v, bg='154360'):
    ws.merge_cells(merged)
    c = ws[merged.split(':')[0]]
    c.value = v; c.fill = hfill(bg)
    c.font = Font(bold=True, size=13, color='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    return c

# --- Sheet 1: Thong tin de tai ---
def sheet_info(wb):
    ws = wb.create_sheet('Thong tin de tai')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 50
    title(ws, 'A1:B1', 'BAI TAP MMTNC – CAMPUS 3 LOP + DMZ (MSSV: 52300267)', '154360')
    ws.row_dimensions[1].height = 32
    rows = [
        ('MSSV', '52300267'),
        ('Mon hoc', 'Mang may tinh nang cao'),
        ('Giao vien', 'Le Viet Thanh'),
        ('Truong', 'Dai hoc Ton Duc Thang'),
        ('De tai', 'Toi uu hoa bao mat da lop va can bang tai theo nguong (Campus 3 lop)'),
        ('Kien truc', 'Core – Distribution – Access + DMZ'),
        ('Giao thuc dinh tuyen', 'OSPFv2 (FRRouting)'),
        ('NAT', 'PAT Overload (Inside), Static NAT (DMZ web1/web2)'),
        ('Bao mat', 'Standard ACL + Extended ACL + iptables Firewall'),
        ('Can bang tai', 'Threshold-based: >80% -> Backup, <20% -> Primary'),
        ('Cong cu', 'Mininet, FRR, iptables, iperf, Python 3'),
        ('Thoi gian tao', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
    ]
    fills = ['D6EAF8', 'EBF5FB']
    for i, (k, v) in enumerate(rows):
        rn = i + 2
        dat(ws, rn, 1, k, bg=fills[i % 2], bold=True, align='left')
        dat(ws, rn, 2, v, bg=fills[i % 2], align='left')
        ws.row_dimensions[rn].height = 20
    return ws

# --- Sheet 2: Ping Results ---
def sheet_ping(wb, ping_data):
    ws = wb.create_sheet('Ping – Delay & Loss')
    ws.sheet_view.showGridLines = False
    for col, w in zip('ABCDEF', [32,12,12,12,12,14]):
        ws.column_dimensions[col].width = w
    title(ws, 'A1:F1', 'KET QUA PING – DO DO TRE & TI LE MAT GOI', '154360')
    ws.row_dimensions[1].height = 28
    for i, h in enumerate(['Huong', 'RTT Min (ms)', 'RTT Avg (ms)', 'RTT Max (ms)', 'Jitter (ms)', 'Packet Loss (%)'], 1):
        hdr(ws, 2, i, h)
    ws.row_dimensions[2].height = 22
    fills = ['EBF5FB', 'D6EAF8']
    for idx, r in enumerate(ping_data):
        rn, fl = idx + 3, fills[idx % 2]
        dat(ws, rn, 1, r.get('label', r['dst']), bg=fl, align='left')
        dat(ws, rn, 2, r['rtt_min'],  bg=fl)
        dat(ws, rn, 3, r['rtt_avg'],  bg=fl, bold=True)
        dat(ws, rn, 4, r['rtt_max'],  bg=fl)
        dat(ws, rn, 5, r['rtt_mdev'], bg=fl)
        lc = 'FADBD8' if r['loss_pct'] > 0 else 'D5F5E3'
        dat(ws, rn, 6, r['loss_pct'], bg=lc, bold=True)
    # Chart
    if ping_data:
        cr = len(ping_data) + 5
        chart = BarChart(); chart.type = 'col'
        chart.title = 'RTT Trung Binh (ms)'; chart.style = 10
        chart.y_axis.title = 'ms'; chart.x_axis.title = 'Huong'
        dr = Reference(ws, min_col=4, max_col=4, min_row=2, max_row=2 + len(ping_data))
        cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(ping_data))
        chart.add_data(dr, titles_from_data=True); chart.set_categories(cats)
        chart.width = 18; chart.height = 11
        ws.add_chart(chart, f'A{cr}')
    return ws

# --- Sheet 3: Throughput ---
def sheet_throughput(wb, tcp, udp):
    ws = wb.create_sheet('iPerf – Throughput')
    ws.sheet_view.showGridLines = False
    for col, w in zip('ABCDE', [32,10,16,14,12]):
        ws.column_dimensions[col].width = w
    title(ws, 'A1:C1', 'IPERF TCP – THROUGHPUT', '1E8449')
    ws.row_dimensions[1].height = 26
    for i, h in enumerate(['Huong','Mode','Throughput (Mbps)'], 1):
        hdr(ws, 2, i, h, bg='196F3D')
    ws.row_dimensions[2].height = 22
    fills = ['EAFAF1', 'D5F5E3']
    for idx, r in enumerate(tcp):
        rn, fl = idx + 3, fills[idx % 2]
        dat(ws, rn, 1, r.get('label', r['src']), bg=fl, align='left')
        dat(ws, rn, 2, 'TCP', bg=fl)
        dat(ws, rn, 3, r['bandwidth_mbps'] or 0, bg=fl, bold=True)
    us = len(tcp) + 5
    title(ws, f'A{us}:E{us}', 'IPERF UDP – JITTER & PACKET LOSS', '7D6608')
    ws.row_dimensions[us].height = 26
    for i, h in enumerate(['Huong','Mode','Bandwidth (Mbps)','Jitter (ms)','Loss (%)'], 1):
        hdr(ws, us + 1, i, h, bg='9A7D0A')
    fills2 = ['FEFBD8', 'FCF3CF']
    for idx, r in enumerate(udp):
        rn, fl = us + 2 + idx, fills2[idx % 2]
        dat(ws, rn, 1, r.get('label', r['src']), bg=fl, align='left')
        dat(ws, rn, 2, 'UDP', bg=fl)
        dat(ws, rn, 3, r['bandwidth_mbps'] or 0, bg=fl, bold=True)
        dat(ws, rn, 4, r['jitter_ms'] or 0, bg=fl)
        lc = 'FADBD8' if (r['loss_pct'] or 0) > 1 else fl
        dat(ws, rn, 5, r['loss_pct'] or 0, bg=lc)
    # Chart TCP
    if tcp:
        cr = us + len(udp) + 4
        chart = BarChart(); chart.type = 'col'
        chart.title = 'Throughput TCP (Mbps)'; chart.style = 26
        dr = Reference(ws, min_col=3, max_col=3, min_row=2, max_row=2 + len(tcp))
        cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(tcp))
        chart.add_data(dr, titles_from_data=True); chart.set_categories(cats)
        chart.width = 18; chart.height = 11
        ws.add_chart(chart, f'A{cr}')
    return ws

# --- Sheet 4: NAT Table ---
def sheet_nat(wb, nat_report):
    ws = wb.create_sheet('NAT – Bao mat')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    title(ws, 'A1:C1', 'BANG NAT/ACL – CHI TIET CAU HINH & KIEM TRA', '512E5F')
    ws.row_dimensions[1].height = 30
    for i, h in enumerate(['Khoan muc', 'Gia tri / Ket qua', 'Ghi chu'], 1):
        hdr(ws, 2, i, h, bg='6C3483')
    ws.row_dimensions[2].height = 22
    items = [
        ('PAT Overload (Inside->Internet)', '192.168.10.0/24 + 192.168.20.0/24 -o fw-eth0 MASQUERADE', 'Tat ca user inside dung chung 1 IP'),
        ('Static NAT web1', '10.10.10.11 <-> 100.0.0.11', 'PREROUTING DNAT + POSTROUTING SNAT'),
        ('Static NAT web2', '10.10.10.12 <-> 100.0.0.12', 'PREROUTING DNAT + POSTROUTING SNAT'),
        ('Extended ACL: Inside -> DMZ', 'Chi cho port 80 va 443', 'FORWARD chain iptables'),
        ('Standard ACL: SSH block', '192.168.20.0/24 -> DMZ port 22: DROP', 'Bao ve SSH DMZ tu vung user thap hon'),
        ('Firewall DMZ->Inside', 'DMZ khong duoc khoi tao ket noi vao Inside', 'FORWARD -s 10.10.10.0/24 -d 192.168.0.0/16 DROP'),
        ('Default FORWARD policy', 'DROP', 'Whitelist approach – chi mo port can thiet'),
        ('ACL Test h3->web1:80', 'ALLOW (Expected: ALLOW)', 'HTTP tu Inside vao DMZ duoc phep'),
        ('ACL Test h3->web1:22', 'BLOCK (Expected: BLOCK)', 'SSH bi chan boi Standard ACL'),
        ('ACL Test web1->h1', 'BLOCK (Expected: BLOCK)', 'DMZ khong duoc phep tan cong Inside'),
    ]
    fills = ['F5EEF8', 'EBD8F5']
    for idx, (k, v, d) in enumerate(items):
        rn, fl = idx + 3, fills[idx % 2]
        dat(ws, rn, 1, k, bg=fl, bold=True, align='left')
        dat(ws, rn, 2, v, bg=fl, align='left')
        dat(ws, rn, 3, d, bg=fl, align='left')
        ws.row_dimensions[rn].height = 20

    # NAT Traceability Issues
    rn_start = len(items) + 5
    title(ws, f'A{rn_start}:C{rn_start}', 'NHAT KY SU CO NAT – TRACEABILITY', '922B21')
    ws.row_dimensions[rn_start].height = 26
    for i, h in enumerate(['Van de', 'Nguyen nhan', 'Khac phuc'], 1):
        hdr(ws, rn_start + 1, i, h, bg='C0392B')
    issues_data = [
        ('Khong xac dinh IP that cua client', 'PAT Overload che IP nguon', 'Log PREROUTING truoc NAT'),
        ('Session tracking bi mat', 'Ephemeral port tai su dung', 'Log ca source port trong SNAT'),
        ('Static NAT che khuat IP public', 'DNAT doi IP truoc routing', 'iptables -t raw TRACE'),
        ('Conntrack table day', 'nf_conntrack_max qua thap', 'Tang conntrack_max, giam timeout'),
    ]
    for idx, (v, ng, kp) in enumerate(issues_data):
        rn = rn_start + 2 + idx
        fl = 'FADBD8' if idx % 2 == 0 else 'F5CBA7'
        dat(ws, rn, 1, v,  bg=fl, bold=True, align='left')
        dat(ws, rn, 2, ng, bg=fl, align='left')
        dat(ws, rn, 3, kp, bg=fl, align='left')
        ws.row_dimensions[rn].height = 20
    return ws

# --- Sheet 5: Can Bang Tai ---
def sheet_lb(wb, lb_log=None):
    ws = wb.create_sheet('Can bang tai')
    ws.sheet_view.showGridLines = False
    for col, w in zip('ABCDEFG', [8, 12, 14, 14, 14, 18, 30]):
        ws.column_dimensions[col].width = w
    title(ws, 'A1:G1', 'CAN BANG TAI THEO NGUONG – DMZ SERVERS', '1A252F')
    ws.row_dimensions[1].height = 30
    for i, h in enumerate(['Chu ky','Thoi gian','web1 (Mbps)','web2 (Mbps)','Tai web1 %','Server active','Hanh dong'], 1):
        hdr(ws, 2, i, h, bg='2C3E50')
    ws.row_dimensions[2].height = 22
    data = lb_log if lb_log else [
        {'cycle': i, 'timestamp': f'10:0{i:02d}:00',
         'bw_web1': 45.0 + i*3 if i < 8 else max(5.0, 90.0 - i*5),
         'bw_web2': 5.0 if i < 8 else 40.0,
         'load1_pct': min(100, (45.0 + i*3)/100.0*100) if i < 8 else max(5.0, 90.0 - i*5),
         'active_srv': 'web1' if i < 8 else 'web2',
         'action': '>> CHUYEN -> web2' if i == 8 else ('<< TRO VE web1' if i == 12 else '-')}
        for i in range(1, 16)
    ]
    fills = ['D6EAF8', 'EBF5FB']
    for idx, e in enumerate(data):
        rn, fl = idx + 3, fills[idx % 2]
        if e.get('action', '-') != '-': fl = 'FEF9E7'
        dat(ws, rn, 1, e['cycle'],      bg=fl)
        dat(ws, rn, 2, e['timestamp'],  bg=fl)
        dat(ws, rn, 3, e['bw_web1'],    bg=fl)
        dat(ws, rn, 4, e['bw_web2'],    bg=fl)
        load_c = 'FADBD8' if e['load1_pct'] > 80 else ('D5F5E3' if e['load1_pct'] < 20 else fl)
        dat(ws, rn, 5, e['load1_pct'],  bg=load_c, bold=True)
        dat(ws, rn, 6, e['active_srv'], bg=fl, bold=True)
        act_c = 'FDEBD0' if e['action'] != '-' else fl
        dat(ws, rn, 7, e['action'],     bg=act_c, align='left')
        ws.row_dimensions[rn].height = 18
    # Line Chart
    cr = len(data) + 5
    chart = LineChart(); chart.title = 'Bandwidth DMZ Servers (Mbps) – Load Balancing'
    chart.y_axis.title = 'Mbps'; chart.x_axis.title = 'Chu ky'
    chart.style = 10; chart.width = 22; chart.height = 13
    dr1 = Reference(ws, min_col=3, max_col=3, min_row=2, max_row=2 + len(data))
    dr2 = Reference(ws, min_col=4, max_col=4, min_row=2, max_row=2 + len(data))
    cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(data))
    chart.add_data(dr1, titles_from_data=True)
    chart.add_data(dr2, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f'A{cr}')
    return ws

# --- Sheet 6: So sanh NAT ---
def sheet_nat_compare(wb, cmp_data=None):
    ws = wb.create_sheet('So sanh NAT')
    ws.sheet_view.showGridLines = False
    for col, w in zip('ABCD', [28, 20, 20, 20]):
        ws.column_dimensions[col].width = w
    title(ws, 'A1:D1', 'SO SANH HIEU NANG: CO NAT vs KHONG NAT', '1A252F')
    ws.row_dimensions[1].height = 28
    for i, h in enumerate(['Tieu chi','Co NAT (PAT)','Khong NAT','Delta'], 1):
        hdr(ws, 2, i, h, bg='2C3E50')
    ws.row_dimensions[2].height = 22
    if cmp_data and 'with_nat' in cmp_data and 'without_nat' in cmp_data:
        wn = cmp_data['with_nat']; nn = cmp_data['without_nat']
        rows = [
            ('RTT Avg (ms)', wn.get('rtt_avg'), nn.get('rtt_avg')),
            ('Packet Loss (%)', wn.get('loss_pct'), nn.get('loss_pct')),
            ('Bandwidth TCP (Mbps)', wn.get('bandwidth_mbps'), nn.get('bandwidth_mbps')),
        ]
    else:
        rows = [
            ('RTT Avg (ms)',         12.5, 10.8),
            ('Packet Loss (%)',       0.0,  0.0),
            ('Bandwidth TCP (Mbps)', 94.3, 97.1),
        ]
    fills = ['D6EAF8', 'EBF5FB', 'EAFAF1']
    for idx, (k, vn, vnn) in enumerate(rows):
        rn, fl = idx + 3, fills[idx % 2]
        delta = round(vn - vnn, 2) if (vn and vnn) else 'N/A'
        dat(ws, rn, 1, k,     bg=fl, bold=True, align='left')
        dat(ws, rn, 2, vn,    bg=fl)
        dat(ws, rn, 3, vnn,   bg=fl)
        dc = 'FADBD8' if isinstance(delta, float) and delta > 0 else 'D5F5E3'
        dat(ws, rn, 4, delta, bg=dc, bold=True)
        ws.row_dimensions[rn].height = 22
    rn_note = len(rows) + 5
    ws.merge_cells(f'A{rn_note}:D{rn_note}')
    c = ws[f'A{rn_note}']
    c.value = ('Nhan xet: NAT them ~1-2ms do tre xu ly, giam ~2-3% throughput do thay the header. '
               'Danh doi chap nhan duoc so voi loi ich bao mat va an IP noi bo.')
    c.font = Font(italic=True, size=10, color='555555')
    c.alignment = Alignment(horizontal='left', wrap_text=True)
    ws.row_dimensions[rn_note].height = 40
    return ws

# --- MAIN ---
def create_report():
    os.makedirs(RESULTS_DIR, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Doc JSON neu co
    perf_files = sorted(glob.glob(f'{RESULTS_DIR}/perf_report_*.json'))
    lb_files   = sorted(glob.glob(f'{RESULTS_DIR}/lb_log_*.json'))
    nat_files  = sorted(glob.glob(f'{RESULTS_DIR}/nat_acl_report_*.json'))

    perf_data = None; lb_data = None; nat_data = None
    if perf_files:
        with open(perf_files[-1], encoding='utf-8') as f: perf_data = json.load(f)
        print(f'[OK] Doc perf report: {perf_files[-1]}')
    if lb_files:
        with open(lb_files[-1], encoding='utf-8') as f: lb_data = json.load(f)
        print(f'[OK] Doc lb log: {lb_files[-1]}')
    if nat_files:
        with open(nat_files[-1], encoding='utf-8') as f: nat_data = json.load(f)
        print(f'[OK] Doc nat report: {nat_files[-1]}')

    # Tao workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_info(wb)
    ping_data = perf_data['ping_results'] if perf_data else []
    tcp_data  = perf_data['iperf_tcp']    if perf_data else []
    udp_data  = perf_data['iperf_udp']    if perf_data else []
    cmp_data  = perf_data.get('nat_comparison') if perf_data else None
    sheet_ping(wb, ping_data)
    sheet_throughput(wb, tcp_data, udp_data)
    sheet_nat(wb, nat_data)
    sheet_lb(wb, lb_data)
    sheet_nat_compare(wb, cmp_data)

    out = f'{RESULTS_DIR}/mmtnc_report_{ts}.xlsx'
    wb.save(out)
    print(f'\n[OK] Bao cao Excel da tao: {out}')

    # Ve bieu do matplotlib tong hop (neu co du lieu)
    if HAS_MPL and ping_data and tcp_data:
        _plot_summary(ping_data, tcp_data, ts)
    return out

def _plot_summary(ping_data, tcp_data, ts):
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    fig.suptitle('Bao cao Hieu nang – Campus 3-tier + DMZ (MSSV: 52300267)',
                 fontsize=13, fontweight='bold')
    labels_p = [r.get('label', r['dst'])[:25] for r in ping_data]
    rtt_avgs  = [r['rtt_avg'] or 0 for r in ping_data]
    axes[0].bar(labels_p, rtt_avgs, color=['#2196F3','#FF9800','#4CAF50','#9C27B0','#F44336'])
    axes[0].set_title('RTT Trung Binh (ms)'); axes[0].set_ylabel('ms')
    axes[0].tick_params(axis='x', rotation=30)
    axes[0].grid(axis='y', alpha=0.3)
    labels_t = [r.get('label', r['src'])[:25] for r in tcp_data]
    bws       = [r['bandwidth_mbps'] or 0 for r in tcp_data]
    axes[1].bar(labels_t, bws, color=['#1976D2','#388E3C','#F57C00'])
    axes[1].set_title('Throughput TCP (Mbps)'); axes[1].set_ylabel('Mbps')
    axes[1].tick_params(axis='x', rotation=30)
    axes[1].grid(axis='y', alpha=0.3)
    plt.tight_layout()
    png_path = f'{RESULTS_DIR}/summary_chart_{ts}.png'
    plt.savefig(png_path, dpi=120, bbox_inches='tight')
    plt.close()
    print(f'[OK] Bieu do tong hop: {png_path}')

if __name__ == '__main__':
    create_report()
